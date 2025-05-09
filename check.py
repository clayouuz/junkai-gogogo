import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def check_file_exists(filename, folder_path):
    """
    检查文件夹中是否存在与文件名完全匹配的文件（加上.pdf后缀）
    """
    full_path = os.path.join(folder_path, filename + '.pdf')
    return os.path.exists(full_path)

# 假设Excel文件名为your_file.xlsx，文件夹路径为your_folder_path
excel_file_path = 'table.xlsx'
folder_path = 'pdfs'

# 读取Excel文件
df = pd.read_excel(excel_file_path)

# 创建一个用于标记的函数
def mark_row(index, color):
    for col in range(1, df.shape[1] + 1):  # 从第一列开始到最后一列
        ws.cell(row=index + 2, column=col).fill = PatternFill(start_color=color,
                                                                end_color=color,
                                                                fill_type="solid")

# 加载工作簿
wb = load_workbook(excel_file_path)
ws = wb.active

# 遍历DataFrame的每一行
for index, row in df.iterrows():
    filename = row['文件名称']
    if not isinstance(filename, str):
        mark_row(index, "FFFF0000")  # 标记为红色
        continue

    if check_file_exists(filename, folder_path):
        # 文件存在，不做处理
        pass
    else:
        # 文件不存在，标记为红色
        mark_row(index, "FFFF0000")  # 标记为红色

# 保存修改后的Excel文件
wb.save(excel_file_path)